import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ------------------------------------------------------
# 1) Load ISCC + Assets File
# ------------------------------------------------------

num_matched = 0

print("Loading files...")

df_iscc = pd.read_excel(
    "out/ISCC_Certificates_29.01.2026_10.23.xlsx",
    sheet_name="Certificate Database"
)

df_assets = pd.read_excel(
    r"C:/Users/tashif.ahmed/OneDrive - Shell/T&S LCF - Analytics, Digital, and Economics - Shared Documents/00. LCF Data Lakehouse/GSTs/GST Assets/00. Golden Source File of Asset Capacities.xlsm",
    sheet_name="GoldenSource"
)

# ------------------------------------------------------
# 2) Create Strong Normalization Function
# ------------------------------------------------------
def make_key(s):
    return str(s).lower().replace(" ", "").replace('"', "").replace("'", "")

print("Normalizing company strings...")

# Normalized ISCC company name
df_iscc["company_norm"] = df_iscc["Company_Name"].apply(make_key)

# Normalized asset file keys
df_assets["prod_norm"]  = df_assets["Company/Producer"].apply(make_key)
df_assets["short_norm"] = df_assets["Company/Producer Short Name"].apply(make_key)

# We need these columns for matching + overwriting
assets_min = df_assets[[
    "Company/Producer",
    "Company/Producer Short Name",
    "City",
    "prod_norm",
    "short_norm"
]].copy()

# ------------------------------------------------------
# 3) Init Match Result Columns
# ------------------------------------------------------
df_iscc["matched"] = False
df_iscc["matched_shortname"] = None
df_iscc["matched_city"] = None

# ------------------------------------------------------
# 4) MATCH COMPANIES (with progress bar)
# ------------------------------------------------------
print("Matching ISCC companies to Assets...")

for idx in tqdm(df_iscc.index, desc="Matching companies", unit="row"):
    norm_name = df_iscc.at[idx, "company_norm"]

    # Try match against full Company/Producer
    full = assets_min[assets_min["prod_norm"] == norm_name]
    if not full.empty:
        df_iscc.at[idx, "matched"] = True
        df_iscc.at[idx, "matched_shortname"] = full.iloc[0]["Company/Producer Short Name"]
        df_iscc.at[idx, "matched_city"] = full.iloc[0]["City"]
        continue

    # Try match against Short Name
    short = assets_min[assets_min["short_norm"] == norm_name]
    if not short.empty:
        df_iscc.at[idx, "matched"] = True
        df_iscc.at[idx, "matched_shortname"] = short.iloc[0]["Company/Producer Short Name"]
        df_iscc.at[idx, "matched_city"] = short.iloc[0]["City"]
        continue

# ------------------------------------------------------
# 5) Apply updates where matched
# ------------------------------------------------------
print("Applying company & city replacements...")

for idx in tqdm(df_iscc.index, desc="Updating rows", unit="row"):
    if df_iscc.at[idx, "matched"]:
        # Overwrite company name with short name
        df_iscc.at[idx, "Company_Name"] = df_iscc.at[idx, "matched_shortname"]
        # Overwrite city using asset file city
        df_iscc.at[idx, "City"] = df_iscc.at[idx, "matched_city"]

# ------------------------------------------------------
# 6) Construct asset_location column
# ------------------------------------------------------
df_iscc["asset_location"] = df_iscc["Company_Name"].astype(str) + " " + df_iscc["City"].astype(str)

# ------------------------------------------------------
# 7) Save Excel before highlighting
# ------------------------------------------------------
output_path = "iscc_company_city_standardised.xlsx"
df_iscc.to_excel(output_path, index=False)

# ------------------------------------------------------
# 8) Excel Highlighting (green for matched rows)
# ------------------------------------------------------
print("Applying green highlighting to matched rows...")

wb = load_workbook(output_path)
ws = wb.active

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

matched_col = df_iscc.columns.get_loc("matched") + 1
asset_loc_col = df_iscc.columns.get_loc("asset_location") + 1

for row in tqdm(range(2, ws.max_row + 1), desc="Highlighting Excel", unit="row"):
    if ws.cell(row=row, column=matched_col).value:  # True
        ws.cell(row=row, column=asset_loc_col).fill = green_fill
        num_matched += 1

wb.save(output_path)

print("\n✔ DONE! Company normalized + matched + city overwritten + asset_location created + highlighted!")
print("Number of matched asset identifiers: " + str(num_matched))