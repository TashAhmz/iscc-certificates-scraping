from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def apply_styles(output_file, worksheet):

    wb = load_workbook(output_file)
    ws = wb[worksheet]
    ws.auto_filter.ref = ws.dimensions  # Makes all columns filterable

    # Define fill colors (ARGB format)
    green_fill  = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
    red_fill    = PatternFill(start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid")

    # Loop through rows in column A
    for row in range(2, ws.max_row + 1):
        cell = ws[f"A{row}"]
        val = (cell.value or "").strip()
        if val == "Valid":
            cell.fill = green_fill
        elif val == "Expired":
            cell.fill = orange_fill
        elif val in ("Suspended", "Terminated", "Withdrawn"):
            cell.fill = red_fill

    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 17

    def colour_columns(columns_to_colour, hex_colour):
        try:   
            fill = PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")
            for col in columns_to_colour:
                col_idx = None
                for cell in ws[1]:
                    if cell.value == col:
                        col_idx = cell.column
                        break
                if col_idx:
                    for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx):  # Include header
                        for cell in row:
                            cell.fill = fill
        except:
            return
    # ISCC derived columns
    colour_columns(["City", "Country", "Company_Name", "Certificate_Class", "Scope_Description", "Processing_Unit_Type_Description", "Latitude", "Longitude", "Value_Changed"], "10CFE7B7")
    # LCF dervied categories
    colour_columns(["Certificate_Type", "Facility_Grouping", "Region", "Sub_Region", "Asset_Identifier", "Match_Found"], "10B7DFE7")

    wb.save(output_file)

    print(f"Successfully applied styles to {output_file}")